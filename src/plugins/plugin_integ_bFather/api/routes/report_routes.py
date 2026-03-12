from io import BytesIO
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from plugins.plugin_integ_bFather.utils.model_loader import (
    BrewFatherRecipe, CalculoPreco, Malte, Lupulo, Levedura
)

report_bp = Blueprint('plugin_reports', __name__)


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _build_xlsx(filename, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for idx, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(r[idx - 1])) if r[idx - 1] is not None else 0 for r in rows)) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@report_bp.route('/relatorios/precos')
@login_required
def relatorio_precos_api():
    receita_id = request.args.get('receita_id', type=int)
    data_inicio = _parse_date(request.args.get('data_inicio'))
    data_fim = _parse_date(request.args.get('data_fim'))

    query = CalculoPreco.query.order_by(CalculoPreco.data_calculo.desc())
    if receita_id:
        query = query.filter(CalculoPreco.receita_id == receita_id)
    if data_inicio:
        query = query.filter(CalculoPreco.data_calculo >= data_inicio)
    if data_fim:
        query = query.filter(CalculoPreco.data_calculo <= data_fim)

    registros = []
    for calc in query.all():
        receita = BrewFatherRecipe.query.get(calc.receita_id) if calc.receita_id else None
        registros.append({
            'id': calc.id,
            'receita_id': calc.receita_id,
            'receita_nome': receita.name if receita else calc.nome_produto,
            'data_calculo': calc.data_calculo.isoformat() if calc.data_calculo else None,
            'custo_total': float(calc.valor_total or 0),
            'custo_por_litro': float(calc.valor_litro_base or 0),
            'quantidade_ml': calc.quantidade_ml,
            'tipo_embalagem': calc.tipo_embalagem,
            'valor_venda_final': float(calc.valor_venda_final or 0),
            'ingredientes_resumo': f"{calc.quantidade_ml} ml / {calc.tipo_embalagem}"
        })

    return jsonify({'success': True, 'registros': registros})


@report_bp.route('/relatorios/precos/exportar')
@login_required
def relatorio_precos_exportar():
    payload = relatorio_precos_api().get_json()
    registros = payload.get('registros', [])
    rows = [[
        item['receita_nome'],
        item['data_calculo'],
        item['custo_total'],
        item['custo_por_litro'],
        item['quantidade_ml'],
        item['tipo_embalagem'],
        item['valor_venda_final']
    ] for item in registros]
    return _build_xlsx(
        f"relatorio_precos_{datetime.now().strftime('%Y%m%d')}.xlsx",
        'Relatorio Precos',
        ['Receita', 'Data do cálculo', 'Custo total', 'Custo por litro', 'Quantidade (ml)', 'Embalagem', 'Preço final'],
        rows
    )


@report_bp.route('/relatorios/ingredientes')
@login_required
def relatorio_ingredientes_api():
    tipo = request.args.get('tipo', 'todos')
    fornecedor = request.args.get('fornecedor', 'todos').strip().lower()

    registros = []

    def add_items(items, tipo_nome, preco_attr, quantidade_label):
        for item in items:
            fabricante = (item.fabricante or '').strip()
            if fornecedor not in ('', 'todos') and fabricante.lower() != fornecedor:
                continue
            preco = getattr(item, preco_attr, 0) or 0
            registros.append({
                'id': item.id,
                'nome': item.nome,
                'tipo': tipo_nome,
                'fornecedor': fabricante,
                'quantidade_referencia': quantidade_label(item),
                'preco_unitario': float(preco),
                'preco_total': float(preco),
                'data_atualizacao': item.data_atualizacao.isoformat() if item.data_atualizacao else None,
            })

    if tipo in ('todos', 'malte'):
        add_items(Malte.query.filter_by(ativo=True).order_by(Malte.nome).all(), 'malte', 'preco_kg', lambda i: '1 kg')
    if tipo in ('todos', 'lupulo'):
        add_items(Lupulo.query.filter_by(ativo=True).order_by(Lupulo.nome).all(), 'lupulo', 'preco_kg', lambda i: '1 kg')
    if tipo in ('todos', 'levedura'):
        add_items(Levedura.query.filter_by(ativo=True).order_by(Levedura.nome).all(), 'levedura', 'preco_unidade', lambda i: '1 un')

    fornecedores = sorted({r['fornecedor'] for r in registros if r['fornecedor']})
    return jsonify({'success': True, 'registros': registros, 'fornecedores': fornecedores})


@report_bp.route('/relatorios/ingredientes/exportar')
@login_required
def relatorio_ingredientes_exportar():
    payload = relatorio_ingredientes_api().get_json()
    registros = payload.get('registros', [])
    rows = [[
        item['nome'],
        item['tipo'],
        item['fornecedor'],
        item['quantidade_referencia'],
        item['preco_unitario'],
        item['preco_total'],
        item['data_atualizacao'],
    ] for item in registros]
    return _build_xlsx(
        f"relatorio_ingredientes_{datetime.now().strftime('%Y%m%d')}.xlsx",
        'Relatorio Ingredientes',
        ['Nome', 'Tipo', 'Fornecedor', 'Quantidade', 'Preço unitário', 'Preço total', 'Atualizado em'],
        rows
    )
