from datetime import datetime
from io import BytesIO

from flask import Blueprint, send_file, jsonify
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from plugins.plugin_integ_bFather.utils.model_loader import Malte, Lupulo, Levedura

upload_bp = Blueprint('plugin_upload', __name__)


def _export_xlsx(filename, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for idx, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(r[idx - 1])) if r[idx - 1] is not None else 0 for r in rows)) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 45)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@upload_bp.route('/upload/exportar/<string:tipo>')
@login_required
def exportar_ingredientes(tipo):
    tipo = (tipo or '').lower()
    data_tag = datetime.now().strftime('%Y%m%d')

    if tipo == 'maltes':
        items = Malte.query.filter_by(ativo=True).order_by(Malte.nome).all()
        rows = [[i.id, i.nome, i.fabricante, i.cor_ebc, i.poder_diastatico, i.rendimento, i.preco_kg, i.tipo] for i in items]
        return _export_xlsx(f'maltes_exportados_{data_tag}.xlsx', 'Maltes', ['ID', 'Nome', 'Fabricante', 'Cor EBC', 'Poder diastático', 'Rendimento', 'Preço/kg', 'Tipo'], rows)

    if tipo in ('lupulo', 'lupulos'):
        items = Lupulo.query.filter_by(ativo=True).order_by(Lupulo.nome).all()
        rows = [[i.id, i.nome, i.fabricante, i.alpha_acidos, i.beta_acidos, i.formato, i.origem, i.preco_kg, i.aroma] for i in items]
        return _export_xlsx(f'lupulos_exportados_{data_tag}.xlsx', 'Lupulos', ['ID', 'Nome', 'Fabricante', 'Alpha ácidos', 'Beta ácidos', 'Formato', 'Origem', 'Preço/kg', 'Aroma'], rows)

    if tipo in ('levedura', 'leveduras'):
        items = Levedura.query.filter_by(ativo=True).order_by(Levedura.nome).all()
        rows = [[i.id, i.nome, i.fabricante, i.formato, i.atenuacao, i.temp_fermentacao, i.preco_unidade, i.floculacao] for i in items]
        return _export_xlsx(f'leveduras_exportadas_{data_tag}.xlsx', 'Leveduras', ['ID', 'Nome', 'Fabricante', 'Formato', 'Atenuação', 'Temperatura fermentação', 'Preço/unidade', 'Floculação'], rows)

    return jsonify({'success': False, 'error': f'Tipo de exportação inválido: {tipo}'}), 404
